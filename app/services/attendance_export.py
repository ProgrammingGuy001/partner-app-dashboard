"""Attendance workbook: one file, one sheet per subject (IPs, supervisors).

Carries the geofence columns the check-in paths record, which is the whole point of
the export — the app never blocks an out-of-fence check-in, so this is where an
out-of-fence day actually gets seen.
"""

import io
from datetime import date, datetime, time as dtime

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.model.admin_attendance import AdminAttendance
from app.model.attendance import DailyAttendance
from app.model.job import Job
from app.model.user import User

IP_HEADERS = [
    "Date",
    "Type",
    "Name",
    "Phone",
    "Job",
    "Manual location",
    "Latitude",
    "Longitude",
    "Distance from site (m)",
    "Inside geofence",
    "Report status",
    "Recorded at",
    "Photo URL",
    "Report URL",
]

SUPERVISOR_HEADERS = [
    "Date",
    "Supervisor",
    "Manual location",
    "Latitude",
    "Longitude",
    "Nearest site",
    "Distance from site (m)",
    "Inside geofence",
    "Notes",
    "Marked at",
    "Photo URL",
]


def _fence_label(within: bool | None) -> str:
    """Blank-ish middle state matters: the site may simply have no pin to measure against."""
    if within is None:
        return "No geofence set"
    return "Inside" if within else "Outside"


def _end_of_day(value: date) -> datetime:
    return datetime.combine(value, dtime(23, 59, 59))


def _write_sheet(sheet, headers: list[str], rows: list[list]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append(row)
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = min(
            max(len(header) + 4, 12), 40
        )
    sheet.freeze_panes = "A2"


def _ip_rows(
    db: Session,
    *,
    visible_phones: list[str] | None,
    job_id: int | None,
    phone: str | None,
    date_from: date | None,
    date_to: date | None,
    phone_to_name: dict,
    report_status,
) -> list[list]:
    query = db.query(DailyAttendance)
    # visible_phones is None for a superadmin (no scoping); an empty list means the
    # caller supervises nobody, which must export nothing rather than everything.
    if visible_phones is not None:
        if not visible_phones:
            return []
        query = query.filter(DailyAttendance.phone.in_(visible_phones))
    if job_id is not None:
        query = query.filter(DailyAttendance.job_id == job_id)
    if phone:
        query = query.filter(DailyAttendance.phone.ilike(f"%{phone.strip()}%"))
    if date_from:
        query = query.filter(DailyAttendance.recorded_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        query = query.filter(DailyAttendance.recorded_at <= _end_of_day(date_to))

    records = query.order_by(DailyAttendance.recorded_at.desc()).all()
    job_ids = {r.job_id for r in records if r.job_id}
    job_names = {
        job.id: job.name
        for job in db.query(Job.id, Job.name).filter(Job.id.in_(job_ids)).all()
    } if job_ids else {}

    rows = []
    for r in records:
        rows.append([
            r.attendance_date.isoformat() if r.attendance_date else "",
            "Check out" if r.attendance_type == "check_out" else "Check in",
            phone_to_name.get(r.phone, ""),
            r.phone,
            job_names.get(r.job_id) or (f"Job #{r.job_id}" if r.job_id else "Independent"),
            r.manual_location or "",
            r.latitude,
            r.longitude,
            r.distance_meters,
            _fence_label(r.within_geofence),
            report_status(r) or "",
            r.recorded_at.isoformat() if r.recorded_at else "",
            r.photo_url or "",
            r.report_document_url or "",
        ])
    return rows


def _supervisor_rows(
    db: Session,
    *,
    admin_id: int | None,
    date_from: date | None,
    date_to: date | None,
) -> list[list]:
    query = db.query(AdminAttendance)
    if admin_id is not None:
        query = query.filter(AdminAttendance.admin_id == admin_id)
    if date_from:
        query = query.filter(AdminAttendance.marked_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        query = query.filter(AdminAttendance.marked_at <= _end_of_day(date_to))

    records = query.order_by(AdminAttendance.marked_at.desc()).all()
    admin_ids = {r.admin_id for r in records}
    admins = {
        row.id: (row.name or row.email)
        for row in db.query(User.id, User.name, User.email).filter(User.id.in_(admin_ids)).all()
    } if admin_ids else {}
    job_ids = {r.matched_job_id for r in records if r.matched_job_id}
    job_names = {
        job.id: job.name
        for job in db.query(Job.id, Job.name).filter(Job.id.in_(job_ids)).all()
    } if job_ids else {}

    rows = []
    for r in records:
        nearest = job_names.get(r.matched_job_id) or (
            f"Job #{r.matched_job_id}" if r.matched_job_id else ""
        )
        rows.append([
            r.marked_at.date().isoformat() if r.marked_at else "",
            admins.get(r.admin_id, f"admin#{r.admin_id}"),
            r.manual_location or "",
            r.latitude,
            r.longitude,
            nearest,
            r.distance_meters,
            _fence_label(r.within_geofence),
            r.notes or "",
            r.marked_at.isoformat() if r.marked_at else "",
            r.photo_url or "",
        ])
    return rows


def build_attendance_workbook(
    db: Session,
    *,
    visible_phones: list[str] | None,
    phone_to_name: dict,
    include_supervisors: bool,
    report_status,
    job_id: int | None = None,
    phone: str | None = None,
    admin_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
) -> bytes:
    """Both sheets in one file. Supervisor rows are superadmin-only, matching the
    permissions on GET /admin/all-attendance — a plain admin gets a header-only sheet."""
    workbook = Workbook()
    ip_sheet = workbook.active
    ip_sheet.title = "IP Attendance"
    _write_sheet(
        ip_sheet,
        IP_HEADERS,
        _ip_rows(
            db,
            visible_phones=visible_phones,
            job_id=job_id,
            phone=phone,
            date_from=date_from,
            date_to=date_to,
            phone_to_name=phone_to_name,
            report_status=report_status,
        ),
    )

    supervisor_sheet = workbook.create_sheet("Supervisor Attendance")
    _write_sheet(
        supervisor_sheet,
        SUPERVISOR_HEADERS,
        _supervisor_rows(db, admin_id=admin_id, date_from=date_from, date_to=date_to)
        if include_supervisors
        else [],
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
