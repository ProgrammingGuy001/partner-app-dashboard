import re
from copy import copy
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pypdf import PdfReader, PdfWriter
from reportlab.lib.utils import simpleSplit
from reportlab.pdfgen import canvas

from app.schemas.job import MonthlyDocumentRequest, ProjectDocumentRequest

TEMPLATE_DIR = Path(__file__).resolve().parents[1] / "templates" / "documents"
NCR_TEMPLATE = TEMPLATE_DIR / "Level 2- NCR.pdf"
MONTHLY_TEMPLATE = TEMPLATE_DIR / "Monthly executed projects.xlsx"
SIGNOFF_VALUE_X = 262


@dataclass(frozen=True)
class GeneratedDocument:
    content: bytes
    content_type: str
    filename: str


def _safe_name(value: str, fallback: str) -> str:
    name = re.sub(r"[^A-Za-z0-9._ -]", "-", value.strip()).strip(" .-")
    return name or fallback


def _draw_whiteout(canv: canvas.Canvas, x: float, y: float, width: float, height: float) -> None:
    canv.saveState()
    canv.setFillColorRGB(1, 1, 1)
    canv.rect(x, y, width, height, fill=1, stroke=0)
    canv.restoreState()


def _draw_wrapped(
    canv: canvas.Canvas,
    text: str,
    x: float,
    y: float,
    width: float,
    height: float,
    *,
    font_size: float = 7.2,
    align: str = "left",
) -> None:
    value = (text or "").strip()
    if not value:
        return
    lines: list[str] = []
    for part in value.splitlines():
        lines.extend(simpleSplit(part, "Helvetica", font_size, max(width - 6, 5)) or [""])
    line_height = font_size + 1.4
    max_lines = max(1, int((height - 4) // line_height))
    lines = lines[:max_lines]
    if len(lines) == max_lines and max_lines and len(value) > sum(map(len, lines)):
        lines[-1] = lines[-1].rstrip(".") + "..."
    canv.setFont("Helvetica", font_size)
    start_y = y + height - font_size - 3
    for index, line in enumerate(lines):
        line_y = start_y - index * line_height
        if align == "center":
            canv.drawCentredString(x + width / 2, line_y, line)
        else:
            canv.drawString(x + 3, line_y, line)


def _merge_overlay(template: Path, draw) -> bytes:
    reader = PdfReader(str(template))
    page = reader.pages[0]
    page_width = float(page.mediabox.width)
    page_height = float(page.mediabox.height)
    overlay = BytesIO()
    canv = canvas.Canvas(overlay, pagesize=(page_width, page_height))
    draw(canv)
    canv.save()
    overlay.seek(0)
    page.merge_page(PdfReader(overlay).pages[0])
    writer = PdfWriter()
    writer.add_page(page)
    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def generate_ncr_document(job, details: ProjectDocumentRequest) -> GeneratedDocument:
    """Fill the supplied NCR and return a downloadable PDF."""

    sales_order = (
        getattr(job, "sales_order", "") or ""
        if details.data_source == "app"
        else details.sales_order
    )
    rows = details.ncr_rows
    if details.no_ncr_reported or not rows:
        rows = [{
            "sales_order": sales_order,
            "part_no": "",
            "part_description": "",
            "qty": "",
            "issue_description": "No NCR reported",
            "disposition": "",
        }]
    else:
        rows = [row.model_dump() for row in rows]

    def draw(canv: canvas.Canvas) -> None:
        document_number = details.ncr_document_number.strip()
        revision = details.ncr_revision.strip()
        if document_number and document_number != "AYENA-QUA-QF-001":
            _draw_whiteout(canv, 426, 809.6, 146, 8.4)
            canv.setFont("Helvetica", 6.96)
            canv.drawCentredString(499.16, 810.84, document_number)
        if revision and revision != "02-06-2026/Rev-00":
            _draw_whiteout(canv, 426, 786.7, 146, 9.4)
            canv.setFont("Helvetica-Bold", 8.04)
            canv.drawCentredString(499.16, 789.84, revision)

        columns = [
            ("sno", 36.24, 38.42, "center"),
            ("sales_order", 75.14, 58.20, "center"),
            ("part_no", 133.82, 66.14, "center"),
            ("part_description", 200.45, 132.86, "left"),
            ("qty", 333.79, 39.48, "center"),
            ("issue_description", 373.75, 114.98, "left"),
            ("disposition", 489.22, 91.22, "left"),
        ]
        row_bottoms = [
            697.18, 669.70, 643.18, 615.70, 588.22, 560.59, 533.11,
            505.63, 478.15, 451.63, 424.15, 396.65, 369.17, 341.57,
            314.09, 286.61, 260.09, 232.61, 205.10, 177.62, 150.14,
        ]
        compact_table = len(rows) < len(row_bottoms)
        if compact_table:
            table_bottom = row_bottoms[len(rows) - 1]
            _draw_whiteout(canv, 35.5, 70, 545.5, table_bottom - 70)
            canv.setStrokeColorRGB(0, 0, 0)
            canv.setLineWidth(0.5)
            canv.line(35.76, table_bottom, 580.92, table_bottom)

            signoffs = [
                ("1.   City Operation In charge -", details.city_operation_in_charge),
                ("2.   Project Marshal -", details.project_marshal),
                ("3.   Customer Quality Lead -", details.customer_quality_lead),
            ]
            for index, (label, value) in enumerate(signoffs):
                y = table_bottom - 35 - index * 14.5
                canv.setFont("Helvetica-Bold", 9)
                canv.drawString(101, y, label)
                if value.strip():
                    canv.setFont("Helvetica", 9)
                    canv.drawString(SIGNOFF_VALUE_X, y, value.strip())

        for index, row in enumerate(rows[:21]):
            values = {"sno": str(index + 1), **row}
            if not values.get("sales_order"):
                values["sales_order"] = sales_order
            for field, x, width, align in columns:
                _draw_wrapped(canv, values.get(field, ""), x, row_bottoms[index], width, 27, align=align)

        if not compact_table:
            for value, y in [
                (details.city_operation_in_charge, 116.6),
                (details.project_marshal, 102.1),
                (details.customer_quality_lead, 87.6),
            ]:
                if value.strip():
                    canv.setFont("Helvetica-Bold", 9)
                    canv.drawString(SIGNOFF_VALUE_X, y, value.strip())

    content = _merge_overlay(NCR_TEMPLATE, draw)
    if details.data_source == "app":
        project_name = job.name or job.customer_name or f"Job {job.id}"
        fallback_name = f"job-{job.id}"
    else:
        project_name = details.project_name
        fallback_name = "manual"
    project = _safe_name(project_name, fallback_name)
    return GeneratedDocument(content, "application/pdf", f"Level 2 NCR - {project}.pdf")


def _set_cell(ws, coordinate: str, value: str, *, horizontal: str = "center") -> None:
    cell = ws[coordinate]
    cell.value = value
    current = copy(cell.alignment)
    cell.alignment = Alignment(
        horizontal=horizontal,
        vertical=current.vertical or "center",
        wrap_text=True,
    )


def _label_value(label: str, value: str) -> str:
    return f"{label}\n{value}" if value else label


def generate_monthly_document(
    details: MonthlyDocumentRequest,
    jobs_by_id: dict[int, object],
) -> GeneratedDocument:
    """Fill the monthly workbook from manual values or selected app jobs."""
    workbook = load_workbook(MONTHLY_TEMPLATE)
    sheet = workbook["Monthly Executed Projects"]
    _set_cell(sheet, "B6", _label_value("Experience Centre Location", details.experience_centre))
    _set_cell(sheet, "F6", _label_value("Ops Manager Name", details.ops_manager))
    _set_cell(sheet, "I6", _label_value("Month'Year", details.month_year))

    for index, project in enumerate(details.projects):
        row = 10 + index * 2
        job = jobs_by_id[project.job_id] if details.data_source == "app" else None
        _set_cell(sheet, f"B{row}", str(index + 1))
        _set_cell(
            sheet,
            f"C{row}",
            (job.name or project.project_name or f"Job {job.id}")
            if job
            else project.project_name,
            horizontal="left",
        )
        _set_cell(sheet, f"D{row}", str(job.id) if job else project.project_id)
        _set_cell(sheet, f"E{row}", project.handover_date)
        _set_cell(sheet, f"F{row}", project.days_taken)
        _set_cell(sheet, f"G{row}", project.learning, horizontal="left")

    _set_cell(sheet, "D26", _label_value("No. of team needed", details.team_needed))
    _set_cell(sheet, "F26", _label_value("No. of Trained team (In-house & Vendor )", details.trained_team))
    _set_cell(
        sheet,
        "I26",
        _label_value(
            "Training Needed for new Vendor( need to be planned between 20th to 5th of next month)",
            details.training_needed,
        ),
    )
    output = BytesIO()
    workbook.save(output)
    filename = f"Monthly executed projects - {details.month_year}.xlsx"
    return GeneratedDocument(
        output.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename,
    )
