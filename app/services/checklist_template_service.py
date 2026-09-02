import re
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.model.job import Checklist, ChecklistItem, JobChecklistItemStatus


CHECKLIST_WORKBOOK = (
    Path(__file__).resolve().parents[1]
    / "templates"
    / "documents"
    / "All Check-list.xlsx"
)
WORKBOOK_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
CHECKLIST_PDF_FILES = {
    "ISM Checklist": "ISM checklist.pdf",
    "Validation Checklist": "validation checklist.pdf",
    "Site Readiness Checksheet": "Site Readiness checklist.pdf",
    "LC Carcass Checklist": "LC Carcass.pdf",
    "Counter Top & Dado Checklist": "Counter_Top_& Dado_Checkist.pdf",
    "UC & Loft Checklist": "UC & Loft cabinat.pdf",
    "Handover Checklist": "Handover checklist.pdf",
}

# name, previous names, sheet, item column, item rows
CHECKLIST_SPECS = (
    ("ISM Checklist", ("ISM Checklist",), "ISM Checklist", 4, ()),
    (
        "Validation Checklist",
        ("Validation Checklist", "Site Validation"),
        "Validation Checklist",
        2,
        (
            15,
            17,
            19,
            21,
            23,
            25,
            30,
            32,
            34,
            36,
            38,
            40,
            42,
            44,
            46,
            49,
            51,
            53,
            56,
            58,
            60,
            63,
            65,
            67,
            69,
        ),
    ),
    (
        "Site Readiness Checksheet",
        ("Site Readiness Checksheet", "Site Readiness"),
        "Site readiness checksheet",
        4,
        (
            10,
            12,
            14,
            16,
            20,
            22,
            24,
            26,
            28,
            30,
            34,
            36,
            38,
            42,
            44,
            46,
            50,
            52,
            54,
            56,
            58,
            63,
            65,
            69,
            71,
        ),
    ),
    (
        "LC Carcass Checklist",
        ("LC Carcass Checklist", "LC Carcass"),
        "LC Carcass Checklist",
        4,
        (10, 12, 14, 16, 20, 24, 26, 28, 33, 35, 37, 39, 43, 45, 47),
    ),
    (
        "Counter Top & Dado Checklist",
        ("Counter Top & Dado Checklist", "Counter Top and Dado Installation"),
        "Counter Top & Dado Checklist",
        4,
        (10, 14, 16, 19, 21, 24, 27, 29, 31),
    ),
    (
        "UC & Loft Checklist",
        ("UC & Loft Checklist", "Upper and Loft Cabinet"),
        "UC & Loft checklist",
        4,
        (10, 12, 14, 17, 19, 21, 23, 25, 27, 29, 31, 33),
    ),
    (
        "Handover Checklist",
        ("Handover Checklist",),
        "Handover checklist",
        4,
        (
            10,
            12,
            14,
            16,
            18,
            20,
            22,
            24,
            26,
            28,
            30,
            32,
            33,
            35,
            37,
            39,
            43,
            45,
            47,
            49,
            51,
            53,
            55,
        ),
    ),
)


def checklist_pdf_template(checklist_name: str) -> Path | None:
    """Return the supplied printable PDF for a canonical or legacy checklist name."""
    wanted = (checklist_name or "").strip().casefold()
    for canonical_name, aliases, *_rest in CHECKLIST_SPECS:
        if wanted not in {name.casefold() for name in aliases}:
            continue
        path = Path(__file__).resolve().parents[2] / CHECKLIST_PDF_FILES[canonical_name]
        return path if path.is_file() else None
    return None


def _clean_cell(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("_", " ")).strip()


def _match_existing_items(
    existing: list[ChecklistItem], texts: list[str]
) -> dict[int, ChecklistItem]:
    def comparable(value: str) -> str:
        return re.sub(
            r"[^a-z0-9]+", " ", value.casefold().replace("then", "than")
        ).strip()

    candidates = []
    for item in existing:
        for index, text in enumerate(texts):
            score = SequenceMatcher(
                None, comparable(item.text), comparable(text)
            ).ratio()
            if score >= 0.70:
                candidates.append((score, abs(item.position - index - 1), item, index))

    matches: dict[int, ChecklistItem] = {}
    used_ids: set[int] = set()
    for _score, _distance, item, index in sorted(
        candidates, key=lambda value: (-value[0], value[1], value[2].id)
    ):
        if index not in matches and item.id not in used_ids:
            matches[index] = item
            used_ids.add(item.id)
    return matches


def sync_checklist_templates(
    db: Session, workbook_path: Path = CHECKLIST_WORKBOOK
) -> None:
    """Replace global checklist templates from the supplied workbook without changing existing item IDs."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for (
            canonical_name,
            aliases,
            sheet_name,
            item_column,
            item_rows,
        ) in CHECKLIST_SPECS:
            checklist = (
                db.query(Checklist)
                .filter(Checklist.name.in_(aliases))
                .order_by(Checklist.id.asc())
                .first()
            )
            if checklist is None:
                checklist = Checklist(name=canonical_name)
                db.add(checklist)
                db.flush()

            checklist.name = canonical_name
            checklist.description = "Download the printable PDF, complete it on site, then upload the finished checklist."

            sheet = workbook[sheet_name]
            texts = [
                _clean_cell(sheet.cell(row, item_column).value) for row in item_rows
            ]
            if any(not text for text in texts):
                raise ValueError(f"Blank checklist item found in {sheet_name}")

            existing = sorted(
                checklist.checklist_items, key=lambda item: (item.position, item.id)
            )
            # ponytail: fuzzy reuse preserves job history across wording edits; add stable workbook keys if imports become user-configurable.
            matches = _match_existing_items(existing, texts)
            matched_ids = {item.id for item in matches.values()}

            # Free every final position before reordering. PostgreSQL checks the
            # unique (checklist_id, position) constraint during each UPDATE.
            temporary_start = (
                max(len(texts), max((item.position for item in existing), default=0))
                + len(existing)
                + 1
            )
            for offset, item in enumerate(existing):
                item.position = temporary_start + offset
            db.flush()

            for index, text in enumerate(texts):
                item = matches.get(index)
                if item is None:
                    item = ChecklistItem(checklist_id=checklist.id)
                    db.add(item)
                item.text = text
                item.position = index + 1

            preserved_position = len(texts) + 1
            for item in existing:
                if item.id in matched_ids:
                    continue
                has_history = (
                    db.query(JobChecklistItemStatus.id)
                    .filter(JobChecklistItemStatus.checklist_item_id == item.id)
                    .first()
                    is not None
                )
                if not has_history:
                    db.delete(item)
                else:
                    item.position = preserved_position
                    preserved_position += 1

        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        workbook.close()
