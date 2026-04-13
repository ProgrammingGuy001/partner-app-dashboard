import io
import logging
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import List

from fastapi import HTTPException
from sqlalchemy.orm import Session, selectinload

from app.model.site_requisite import SiteRequisite
from app.model.so_detail import SODetail
from app.schemas.requisite_schema import SiteRequisiteSubmit

logger = logging.getLogger(__name__)


class RequisiteService:
    TEMPLATE_CANDIDATES = (
        "repairorder.xlsx",
        "Repair Order (repair.order) (1).xlsx",
    )

    @staticmethod
    def _format_order_status(raw_status: str | None) -> str:
        if not raw_status:
            return ""

        status_map = {
            "draft": "Quotation",
            "sent": "Quotation Sent",
            "sale": "Confirmed",
            "done": "Locked",
            "cancel": "Cancelled",
        }
        normalized = raw_status.strip().lower()
        return status_map.get(normalized, normalized.replace("_", " ").title())

    @staticmethod
    def _normalize_header(header: str | None) -> str:
        return "".join(ch.lower() for ch in (header or "") if ch.isalnum())

    @classmethod
    def _resolve_template_path(cls) -> Path:
        root = Path(__file__).resolve().parents[2]

        for candidate in cls.TEMPLATE_CANDIDATES:
            candidate_path = root / candidate
            if candidate_path.exists():
                return candidate_path

        for path in root.glob("*.xlsx"):
            normalized = path.name.lower().replace(" ", "")
            if "repairorder" in normalized or "repair.order" in normalized:
                return path

        raise FileNotFoundError("Repair order template not found in project root")

    @staticmethod
    def submit_site_requisite(
        db: Session,
        data: SiteRequisiteSubmit,
        user_id: int | None = None,
    ):
        """Submit site requisite with all bucket items."""
        odoo: dict = {}
        try:
            from app.services.odoo_service import OdooService
            odoo = OdooService.get_sales_order_details(data.sales_order)
        except Exception as exc:
            logger.warning("Could not fetch Odoo details for %s: %s", data.sales_order, exc)

        addr_parts = [
            odoo.get("address_line_1", ""),
            odoo.get("address_line_2", ""),
            odoo.get("city", ""),
            odoo.get("state", ""),
            odoo.get("pincode", ""),
        ]
        delivery_address = ", ".join(part for part in addr_parts if part)
        so_status = RequisiteService._format_order_status(odoo.get("order_state"))

        # Always create a new requisite record — one SO can have multiple submissions
        so_detail = SODetail(
            sales_order=data.sales_order,
            sr_poc=data.sr_poc,
            status="pending",
            cabinet_position=data.cabinet_position,
            ip_user_id=user_id,
            customer_name=odoo.get("customer_name", ""),
            project_name=odoo.get("project_name", ""),
            delivery_address=delivery_address,
            so_poc=odoo.get("client_order_ref", ""),
            so_status=so_status,
            repair_reference=data.repair_reference,
            expected_delivery=data.expected_delivery,
            do_number=data.do_number,
        )
        db.add(so_detail)
        db.flush()

        for item in data.items:
            db.add(
                SiteRequisite(
                    so_detail_id=so_detail.id,
                    product_name=item.product_name,
                    quantity=item.quantity,
                    issue_description=item.issue_description,
                    responsible_department=item.responsible_department,
                    component_status=item.component_status,
                )
            )

        db.commit()

        return (
            db.query(SODetail)
            .options(selectinload(SODetail.site_requisites))
            .filter(SODetail.id == so_detail.id)
            .first()
        )

    @staticmethod
    def get_history(
        db: Session,
        limit: int = 50,
        offset: int = 0,
        user_id: int | None = None,
    ) -> List[SODetail]:
        """Get site requisite history. Scoped to user_id when provided."""
        query = db.query(SODetail).options(selectinload(SODetail.site_requisites))
        if user_id is not None:
            query = query.filter(SODetail.ip_user_id == user_id)
        return query.order_by(SODetail.created_date.desc()).offset(offset).limit(limit).all()

    @staticmethod
    def get_history_by_sales_order(
        db: Session,
        sales_order: str,
        user_id: int | None = None,
    ) -> List[SODetail]:
        """Get all requisites for a sales order. Enforces ownership when user_id is provided."""
        query = (
            db.query(SODetail)
            .options(selectinload(SODetail.site_requisites))
            .filter(SODetail.sales_order == sales_order)
        )
        if user_id is not None:
            query = query.filter(SODetail.ip_user_id == user_id)
        return query.order_by(SODetail.created_date.desc()).all()

    @staticmethod
    def update_status(db: Session, so_id: int, status: str):
        """Update SO status."""
        so_detail = db.query(SODetail).filter(SODetail.id == so_id).first()
        if so_detail:
            so_detail.status = status
            if status == "completed":
                so_detail.closed_date = datetime.utcnow()
            db.commit()
            db.refresh(so_detail)
            return so_detail
        return None

    @classmethod
    def generate_repair_order_xlsx(
        cls,
        db: Session,
        so_id: int,
        user_id: int | None = None,
    ) -> bytes:
        """Generate a Repair Order xlsx for a specific requisite record."""
        from openpyxl import load_workbook

        query = (
            db.query(SODetail)
            .options(selectinload(SODetail.site_requisites))
            .filter(SODetail.id == so_id)
        )
        if user_id is not None:
            query = query.filter(SODetail.ip_user_id == user_id)
        so_detail = query.first()
        if not so_detail:
            raise HTTPException(status_code=404, detail="Requisite not found")
        sales_order = so_detail.sales_order

        workbook = load_workbook(cls._resolve_template_path())
        sheet = workbook.active

        template_row = 2 if sheet.max_row >= 2 else 1
        template_styles = {
            column_index: copy(sheet.cell(row=template_row, column=column_index)._style)
            for column_index in range(1, sheet.max_column + 1)
        }
        template_height = sheet.row_dimensions[template_row].height

        if sheet.max_row >= 2:
            sheet.delete_rows(2, sheet.max_row - 1)

        customer_name = so_detail.customer_name or ""
        project_name = so_detail.project_name or ""
        delivery_address = so_detail.delivery_address or ""
        so_poc = so_detail.so_poc or ""
        sr_poc = so_detail.sr_poc or ""
        so_status = so_detail.so_status or so_detail.status.capitalize()
        repair_reference = so_detail.repair_reference or ""
        expected_delivery = (
            so_detail.expected_delivery.strftime("%d-%m-%Y")
            if so_detail.expected_delivery
            else ""
        )
        do_number = so_detail.do_number or ""

        requisites = so_detail.site_requisites or []
        rows = requisites if requisites else [None]

        for row_index, requisite in enumerate(rows, start=2):
            is_first_row = row_index == 2
            line_item_values = {
                "partsproduct": requisite.product_name if requisite else "",
                "product": requisite.product_name if requisite else "",
                "partsdemand": (
                    float(requisite.quantity)
                    if requisite and requisite.quantity is not None
                    else ""
                ),
                "quantity": (
                    float(requisite.quantity)
                    if requisite and requisite.quantity is not None
                    else ""
                ),
                "componentstatus": (
                    (requisite.component_status or "")
                    if requisite
                    else ""
                ),
            }
            first_row_values = {
                "customer": customer_name,
                "customername": customer_name,
                "projectname": project_name,
                "repairreference": repair_reference,
                "status": so_status,
                "saleorder": sales_order,
                "salesorder": sales_order,
                "sopoc": so_poc,
                "expecteddelivery": expected_delivery,
                "deliveryaddress": delivery_address,
                "srpoc": sr_poc,
                "donumber": do_number,
            }

            for column_index in range(1, sheet.max_column + 1):
                header_key = cls._normalize_header(sheet.cell(row=1, column=column_index).value)
                if header_key in line_item_values:
                    value = line_item_values[header_key]
                elif is_first_row:
                    value = first_row_values.get(header_key, "")
                else:
                    value = ""

                cell = sheet.cell(row=row_index, column=column_index, value=value)
                cell._style = copy(template_styles[column_index])

            if template_height is not None:
                sheet.row_dimensions[row_index].height = template_height

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.read()
